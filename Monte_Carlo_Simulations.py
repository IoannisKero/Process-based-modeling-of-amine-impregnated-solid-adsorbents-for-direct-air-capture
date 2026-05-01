# Copyright (c) 2026 Ioannis Keroglou. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the LICENSE file for the specific language governing permissions
# and limitations under the License.
"""
Author: Ioannis Keroglou
Affiliation: School of Sustainability Engineering and Environmental Engineering,
             Purdue University, West Lafayette, IN, USA 
Date: April 2026
Project: Process-based modelling of direct air capture sorbent production:
         A comparative life cycle assessment of amine-impregnated solid adsorbents
"""

# Import Libraries
import pandas as pd
import numpy as np

EXCLUDED_MCS_SORBENT_TOKENS = ()


def _is_excluded_mcs_sorbent(sorbent_name):
    s = str(sorbent_name).upper()
    return any(tok in s for tok in EXCLUDED_MCS_SORBENT_TOKENS)


def _filter_excluded_mcs_sorbents_df(df):
    """Remove excluded sorbent types from MCS output dataframes before saving."""
    if df is None or df.empty:
        return df
    for col in ("Sorbent Type", "Sorbent", "Sorbent_Type"):
        if col in df.columns:
            return df[~df[col].astype(str).apply(_is_excluded_mcs_sorbent)].reset_index(drop=True)
    return df


# ========================================================================================================================
# STEP 1: Monte Carlo Simulations - Mass Dictionaries
# ========================================================================================================================

def monte_carlo_statistics_mass(mass_dictionary, adsorption_performance_dictionary, stage_name, scenarios):
    """
    Calculate Monte Carlo statistics for 3D mass matrices.
    
    Parameters:
    -----------
    mass_dictionary : dict
        Nested dictionary with structure: {scenario_id: {sorbent_type: {component: 3D_matrix}}}
        where 3D_matrix has shape (1, steps, mcs_number)
    component_name : str
        Name for the component (e.g., "Aziridine", "PEI", "Support", "Sorbent")
    adsorption_performance_dictionary : dict
        Dictionary containing adsorption performance results from adsorption_performance()
        Structure: {sorbent_type: {"N": array, "k": array, "loadings": {loading: {condition: array}}}}
    scenarios : dict
        Scenarios dictionary containing batch_size and pei_composition information
        
    Returns:
    --------
    pd.DataFrame
        DataFrame with statistics for each scenario/sorbent/component/step combination
    """
    
    summary_list = []
    
    # Iterate through scenarios
    for scenario_id, scenario_data in mass_dictionary.items():
        # Get scenario information
        batch_size_kg = scenarios[scenario_id]['profile']['batch_size']/1000  # Convert to kg
        pei_wt = scenarios[scenario_id]['profile']['pei_composition']
        
        # Iterate through sorbent types
        for sorbent_type, sorbent_data in scenario_data.items():
            # Skip excluded sorbent types in MCS outputs
            if _is_excluded_mcs_sorbent(sorbent_type):
                continue
            # Iterate through components
            for component, matrix in sorbent_data.items():
                if isinstance(matrix, np.ndarray) and matrix.ndim == 3:
                    # Calculate statistics for each step separately
                    for step in range(matrix.shape[1]):  # Iterate through each step
                        # Extract MCS values for this specific step
                        mcs_values = matrix[0, step, :]  
                        num_samples = len(mcs_values)
                        
                        # Calculate statistics for this step
                        mean_val = np.mean(mcs_values)
                        std_val = np.std(mcs_values)
                        sem_val = std_val / np.sqrt(num_samples)
                        sem_percent = (sem_val / mean_val) * 100 if mean_val != 0 else np.nan
                        ci_lower = np.percentile(mcs_values, 2.5)
                        ci_upper = np.percentile(mcs_values, 97.5)

                        summary_list.append({
                            'Stage': stage_name,
                                    'Scenario': scenario_id,
                                    'Sorbent Type': sorbent_type,
                                    'Component': component,
                                    'Step': step,
                                    'PEI wt%': pei_wt,
                                    'Batch Size (kg)': batch_size_kg,
                                    '95% CI Lower (kg/kg sorbent)': (ci_lower/1000)/batch_size_kg,
                                    'Mean (kg/kg sorbent)': (mean_val/1000)/batch_size_kg,
                                    '95% CI Upper (kg/kg sorbent)': (ci_upper/1000)/batch_size_kg,
                                    'Sample Size': num_samples,
                                })

    # Create DataFrame with all statistics
    MC_stats_df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))
    
    # Save to Excel file (create new or append to existing)
    try:
        # Try to append to existing file
        with pd.ExcelWriter('Life_Cycle_Inventory.xlsx', engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sheet_name = f"{stage_name}_Mass_Stats"
            MC_stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Saved {len(MC_stats_df)} rows to {sheet_name} sheet in Life_Cycle_Inventory.xlsx")
    except FileNotFoundError:
        # Create new file if it doesn't exist
        with pd.ExcelWriter('Life_Cycle_Inventory.xlsx', engine="openpyxl") as writer:
            sheet_name = f"{stage_name}_Mass_Stats"
            MC_stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Created new file and saved {len(MC_stats_df)} rows to {sheet_name} sheet in Life_Cycle_Inventory.xlsx")
    
    return MC_stats_df

# ========================================================================================================================
# STEP 2: Monte Carlo Simulations - Compound Input Mass per stage dictionary 
# ========================================================================================================================

def monte_carlo_statistics_input_mass(compound_input_mass_dictionary, adsorption_performance_dictionary, scenarios, stage_name="Input"):
    """
    Compute MC statistics for input masses from structured dictionaries produced by material_mass_input.

    Expects compound_input_mass_dictionary shaped as:
      {scenario: {sorbent_type: {stage_name: {component: {mcs_index: {"Input Mass (kg)", "Input Mass (kg/kg product)"}}}}}}

    Returns a DataFrame with stats for both kg and kg/kg product, and kg/kg CO2 captured (Sorbent stage only).
    """

    summary_list = []
    # Column label for normalized values coming from material_mass_input (kg/kg product)
    product_unit = "kg product"

    for scenario_id, sorbents_map in compound_input_mass_dictionary.items():
        # scenario metadata
        batch_size_kg = scenarios[scenario_id]['profile']['batch_size']/1000.0
        pei_wt = scenarios[scenario_id]['profile']['pei_composition']

        for sorbent_type, stages_map in sorbents_map.items():
            # Skip excluded sorbent types in MCS outputs
            if _is_excluded_mcs_sorbent(sorbent_type):
                continue
            stage_block = stages_map.get(stage_name)
            if not isinstance(stage_block, dict):
                continue

            for component, mcs_map in stage_block.items():
                if not isinstance(mcs_map, dict):
                    continue

                # Build vectors from per-MCS rows
                mcs_indices = sorted(mcs_map.keys())
                kg_vec = np.array([mcs_map[i]["Input Mass (kg)"] for i in mcs_indices], dtype=float)
                kg_per_kg_sorbent = kg_vec / batch_size_kg
                norm_vec = np.array([
                    mcs_map[i].get("Input Mass (kg/kg product)", np.nan) for i in mcs_indices
                ], dtype=float)
                n = len(kg_vec)
                if n == 0:
                    continue

                # Stats helpers
                def stats(v):
                    return (
                        float(np.mean(v)),
                        float(np.std(v)),
                        float(np.percentile(v, 2.5)),
                        float(np.percentile(v, 97.5)),
                    )

                mean_kg, std_kg, lo_kg, hi_kg = stats(kg_vec)
                mean_kg_per_kg_sorbent, std_kg_per_kg_sorbent, lo_kg_per_kg_sorbent, hi_kg_per_kg_sorbent = stats(kg_per_kg_sorbent)
                mean_norm, std_norm, lo_norm, hi_norm = stats(norm_vec)

                # Calculate per kg CO2 captured stats (Sorbent stage only) using normalized vector
                pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                loading = pei_to_loading_map.get(pei_wt, pei_wt)
                for condition in ["Dry", "Humid"]:
                    mean_capt = std_capt = lo_capt = hi_capt = np.nan
                    if sorbent_type in adsorption_performance_dictionary:
                        sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                        if (
                            isinstance(sorbent_perf, dict)
                            and "loadings" in sorbent_perf
                            and loading in sorbent_perf["loadings"]
                            and condition in sorbent_perf["loadings"][loading]
                        ):
                            perf = np.asarray(sorbent_perf["loadings"][loading][condition], dtype=float)
                            if perf.shape[0] == n:
                                with np.errstate(divide='ignore', invalid='ignore'):
                                    per_capt = np.divide(kg_per_kg_sorbent, perf, out=np.full_like(kg_per_kg_sorbent, np.nan), where=perf>0)
                                mean_capt = float(np.mean(per_capt))
                                std_capt = float(np.std(per_capt))
                                lo_capt = float(np.percentile(per_capt, 2.5))
                                hi_capt = float(np.percentile(per_capt, 97.5))

                    summary_list.append({
                        "Scenario": scenario_id,
                        "Sorbent Type": sorbent_type,
                        "Stage": stage_name,
                        "Component": component,
                        "PEI wt%": pei_wt,
                        "Batch Size (kg)": batch_size_kg,
                        "Condition": condition,
                        "95% CI Lower (kg)": lo_kg,
                        "Mean (kg)": mean_kg,
                        "95% CI Upper (kg)": hi_kg,
                        "95% CI Lower (kg/kg sorbent)": lo_kg_per_kg_sorbent,
                        "Mean (kg/kg sorbent)": mean_kg_per_kg_sorbent,
                        "95% CI Upper (kg/kg sorbent)": hi_kg_per_kg_sorbent,
                        f"95% CI Lower (kg/{product_unit})": lo_norm,
                        f"Mean (kg/{product_unit})": mean_norm,
                        f"95% CI Upper (kg/{product_unit})": hi_norm,
                        "95% CI Lower (kg/kg CO2 captured)": lo_capt,
                        "Mean (kg/kg CO2 captured)": mean_capt,
                        "95% CI Upper (kg/kg CO2 captured)": hi_capt,
                        "Sample Size": n
                    })
    
    # Create DataFrame with all statistics
    MC_stats_df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))
    
    # Save to Excel file (create new or append to existing)
    try:
        # Try to append to existing file
        with pd.ExcelWriter('Life_Cycle_Inventory.xlsx', engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sheet_name = f"{stage_name}_Input_Mass"
            MC_stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Saved {len(MC_stats_df)} rows to {sheet_name} sheet in Life_Cycle_Inventory.xlsx")
    except FileNotFoundError:
        # Create new file if it doesn't exist
        with pd.ExcelWriter('Life_Cycle_Inventory.xlsx', engine="openpyxl") as writer:
            sheet_name = f"{stage_name}_Input_Mass"
            MC_stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Created new file and saved {len(MC_stats_df)} rows to {sheet_name} sheet in Life_Cycle_Inventory.xlsx")
    
    return MC_stats_df


# ========================================================================================================================
# STEP 3: Monte Carlo Simulations - Energy & Material Impact Dictionaries
# ========================================================================================================================

def monte_carlo_statistics_energy_impacts(energy_impact_dictionary, adsorption_performance_dictionary, scenarios):
    """
    Compute MC statistics for energy impact dictionaries.

    Expects energy_impact_dictionary shaped as:
      {scenario: {sorbent: {region_or_heat: {stage: {process/Stage_Total: {impact_category: matrix}},
                                            "Total": {impact_category: matrix}}}}}

    """
    summary_list = []

    for scenario_id, sorbents_map in energy_impact_dictionary.items():
        batch_size_kg = scenarios[scenario_id]["profile"]["batch_size"] / 1000.0
        pei_wt = scenarios[scenario_id]["profile"]["pei_composition"]

        for sorbent_type, regions_map in sorbents_map.items():
            # Skip excluded sorbent types in MCS outputs
            if _is_excluded_mcs_sorbent(sorbent_type):
                continue
            # Track Stage_Total matrices for grand total calculation (only for Direct)
            stage_total_matrices_by_region = {}  # {region: {stage: {category: matrix}}}
            for region_name, region_data in regions_map.items():
                # Process stage-level blocks only here.
                # Grand "Total" is appended once later from summed Stage_Total matrices.
                stages_to_process = []
                # Add all individual stages
                for stage_key in region_data.keys():
                    if stage_key not in {"Total", "Stage_Total"} and isinstance(region_data[stage_key], dict):
                        stage_block = region_data[stage_key]
                        # Add Stage_Total if it exists
                        if "Stage_Total" in stage_block:
                            stages_to_process.append((stage_key, "Stage_Total"))
                        # Process all individual processes within each stage
                        for process_key in stage_block.keys():
                            if process_key != "Stage_Total":
                                process_block = stage_block.get(process_key)
                                # Check if it's a dict (containing impact categories) or a valid matrix
                                if isinstance(process_block, dict) and len(process_block) > 0:
                                    stages_to_process.append((stage_key, process_key))
                
                # Process each stage/process combination
                for current_stage, process_name in stages_to_process:
                    if current_stage == "Total":
                        impact_block = region_data.get("Total")
                    else:
                        stage_block = region_data.get(current_stage, {})
                        if isinstance(stage_block, dict):
                            impact_block = stage_block.get(process_name) if process_name in stage_block else None
                        else:
                            impact_block = None

                    if not isinstance(impact_block, dict):
                        continue

                    for category, matrix in impact_block.items():
                        # Store Stage_Total matrices for grand total calculation
                        if process_name == "Stage_Total" and current_stage != "Total":
                            if region_name not in stage_total_matrices_by_region:
                                stage_total_matrices_by_region[region_name] = {}
                            if current_stage not in stage_total_matrices_by_region[region_name]:
                                stage_total_matrices_by_region[region_name][current_stage] = {}
                            stage_total_matrices_by_region[region_name][current_stage][category] = matrix
                        if not (isinstance(matrix, np.ndarray) and matrix.ndim == 3):
                            continue

                        mcs_values = matrix[0, 0, :]
                        n = len(mcs_values)
                        if n == 0:
                            continue

                        ddof = 1 if n > 1 else 0
                        mean_val = float(np.mean(mcs_values))
                        std_val = float(np.std(mcs_values, ddof=ddof))
                        sem_val = std_val / np.sqrt(n) if n > 0 else np.nan
                        sem_percent = (sem_val / mean_val) * 100 if mean_val != 0 else np.nan
                        ci_lower = float(np.percentile(mcs_values, 2.5))
                        ci_upper = float(np.percentile(mcs_values, 97.5))

                        # Energy impact matrices are already normalized per kg sorbent.
                        per_kg_sorbent_values = mcs_values.copy()
                        per_kg_sorbent_lower = float(np.percentile(per_kg_sorbent_values, 2.5))
                        per_kg_sorbent_mean = float(np.mean(per_kg_sorbent_values))
                        per_kg_sorbent_upper = float(np.percentile(per_kg_sorbent_values, 97.5))

                        pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                        loading = pei_to_loading_map.get(pei_wt, pei_wt)

                        for condition in ["Dry", "Humid"]:
                            per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                            if sorbent_type in adsorption_performance_dictionary:
                                sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                                if (
                                    isinstance(sorbent_perf, dict)
                                    and "loadings" in sorbent_perf
                                    and loading in sorbent_perf["loadings"]
                                    and condition in sorbent_perf["loadings"][loading]
                                ):
                                    perf = np.asarray(sorbent_perf["loadings"][loading][condition], dtype=float)
                                    if perf.shape[0] == n:
                                        # Per kg CO2 captured = per kg sorbent / performance (kg CO2/kg sorbent)
                                        with np.errstate(divide='ignore', invalid='ignore'):
                                            per_capt = np.divide(per_kg_sorbent_values, perf, out=np.full_like(per_kg_sorbent_values, np.nan), where=perf > 0)
                                        per_kg_captured_lower = float(np.percentile(per_capt, 2.5))
                                        per_kg_captured_mean = float(np.mean(per_capt))
                                        per_kg_captured_upper = float(np.percentile(per_capt, 97.5))

                            summary_list.append({
                            "Scenario": scenario_id,
                            "Sorbent Type": sorbent_type,
                            "Region": region_name,
                            "Stage": current_stage,
                            "Process": process_name,
                            "Impact Category": category,
                            "PEI wt%": pei_wt,
                            "Batch Size (kg)": batch_size_kg,
                            "Condition": condition,
                            "95% CI Lower": ci_lower,
                            "Mean": mean_val,
                            "95% CI Upper": ci_upper,
                            "95% CI Lower (per kg sorbent)": per_kg_sorbent_lower,
                            "Mean (per kg sorbent)": per_kg_sorbent_mean,
                            "95% CI Upper (per kg sorbent)": per_kg_sorbent_upper,
                            "95% CI Lower (per kg CO2 captured)": per_kg_captured_lower,
                            "Mean (per kg CO2 captured)": per_kg_captured_mean,
                            "95% CI Upper (per kg CO2 captured)": per_kg_captured_upper,
                            "Sample Size": n,
                        })
            
            # Calculate grand total across all stages for energy impacts
            if stage_total_matrices_by_region:
                for region_name, stages_dict in stage_total_matrices_by_region.items():
                    # Get all impact categories from any stage
                    all_categories = set()
                    for stage_dict in stages_dict.values():
                        all_categories.update(stage_dict.keys())
                    
                    # Calculate grand total for each impact category
                    for category in all_categories:
                        grand_total_matrix = None
                        reference_shape = None
                        
                        # Get reference shape from first available matrix
                        for stage, category_dict in stages_dict.items():
                            if category in category_dict:
                                matrix = category_dict[category]
                                if matrix is not None:
                                    reference_shape = matrix.shape
                                    grand_total_matrix = np.zeros(reference_shape)
                                    break
                        
                        # Sum all Stage_Total matrices for this category across all stages
                        if grand_total_matrix is not None:
                            for stage, category_dict in stages_dict.items():
                                if category in category_dict:
                                    matrix = category_dict[category]
                                    if matrix is not None and matrix.shape == reference_shape:
                                        grand_total_matrix += matrix
                                    elif matrix is not None:
                                        grand_total_matrix += np.broadcast_to(matrix, reference_shape)
                            
                            # Calculate statistics for grand total
                            mcs_values = grand_total_matrix[0, 0, :]
                            n = len(mcs_values)
                            if n > 0:
                                ddof = 1 if n > 1 else 0
                                mean_val = float(np.mean(mcs_values))
                                ci_lower = float(np.percentile(mcs_values, 2.5))
                                ci_upper = float(np.percentile(mcs_values, 97.5))
                                
                                # Energy impact matrices are already normalized per kg sorbent.
                                per_kg_sorbent_values = mcs_values.copy()
                                per_kg_sorbent_lower = float(np.percentile(per_kg_sorbent_values, 2.5))
                                per_kg_sorbent_mean = float(np.mean(per_kg_sorbent_values))
                                per_kg_sorbent_upper = float(np.percentile(per_kg_sorbent_values, 97.5))
                                
                                pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                                loading = pei_to_loading_map.get(pei_wt, pei_wt)
                                
                                for condition in ["Dry", "Humid"]:
                                    per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                                    if sorbent_type in adsorption_performance_dictionary:
                                        sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                                        if (
                                            isinstance(sorbent_perf, dict)
                                            and "loadings" in sorbent_perf
                                            and loading in sorbent_perf["loadings"]
                                            and condition in sorbent_perf["loadings"][loading]
                                        ):
                                            perf = np.asarray(sorbent_perf["loadings"][loading][condition], dtype=float)
                                            if perf.shape[0] == n:
                                                # Per kg CO2 captured = per kg sorbent / performance (kg CO2/kg sorbent)
                                                with np.errstate(divide='ignore', invalid='ignore'):
                                                    per_capt = np.divide(per_kg_sorbent_values, perf, out=np.full_like(per_kg_sorbent_values, np.nan), where=perf > 0)
                                                per_kg_captured_lower = float(np.percentile(per_capt, 2.5))
                                                per_kg_captured_mean = float(np.mean(per_capt))
                                                per_kg_captured_upper = float(np.percentile(per_capt, 97.5))
                                
                                    summary_list.append({
                                        "Scenario": scenario_id,
                                        "Sorbent Type": sorbent_type,
                                        "Region": region_name,
                                        "Stage": "Total",
                                        "Process": "Total",
                                        "Impact Category": category,
                                        "PEI wt%": pei_wt,
                                        "Batch Size (kg)": batch_size_kg,
                                        "Condition": condition,
                                        "95% CI Lower": ci_lower,
                                        "Mean": mean_val,
                                        "95% CI Upper": ci_upper,
                                        "95% CI Lower (per kg sorbent)": per_kg_sorbent_lower,
                                        "Mean (per kg sorbent)": per_kg_sorbent_mean,
                                        "95% CI Upper (per kg sorbent)": per_kg_sorbent_upper,
                            "95% CI Lower (per kg CO2 captured)": per_kg_captured_lower,
                            "Mean (per kg CO2 captured)": per_kg_captured_mean,
                            "95% CI Upper (per kg CO2 captured)": per_kg_captured_upper,
                            "Sample Size": n,
                        })

    def _safe_sheet_name(name):
        safe = name.replace(" ", "_")
        return safe[:31]
    df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))
    # Remove any rows where Stage is a Stage_Total summary label.
    if "Stage" in df.columns:
        df = df[~df["Stage"].astype(str).str.contains("Stage_Tota", case=False, na=False)].reset_index(drop=True)
    # Enforce one row per logical key (prevents duplicated totals/components in saved sheet).
    if not df.empty:
        dedup_keys = [
            "Scenario", "Sorbent Type", "Region", "Stage", "Process",
            "Impact Category", "PEI wt%", "Batch Size (kg)", "Condition"
        ]
        dedup_keys = [k for k in dedup_keys if k in df.columns]
        if dedup_keys:
            before = len(df)
            df = df.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)
            removed = before - len(df)
            if removed > 0:
                print(f"ℹ️ Removed {removed} duplicate material-impact rows before saving.")
    sheet_name = _safe_sheet_name(f"LC_Energy_Impacts")
    try:
        with pd.ExcelWriter("Total_Environmental_Impacts.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Saved {len(df)} rows to sheet '{sheet_name}' in Total_Environmental_Impacts.xlsx")
    except FileNotFoundError:
        with pd.ExcelWriter("Total_Environmental_Impacts.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Created file and saved {len(df)} rows to sheet '{sheet_name}' in Total_Environmental_Impacts.xlsx")

    return df


def monte_carlo_statistics_material_impacts(material_impact_dictionary, adsorption_performance_dictionary, scenarios):
    """
    Compute MC statistics for material impact dictionaries.

    Expects material_impact_dictionary shaped as:
      {scenario: {sorbent: {stage: {component: {impact_category: matrix}, "Stage_Total": {...}}, "Total": {...}}}}
    """
    summary_list = []

    for scenario_id, sorbents_map in material_impact_dictionary.items():
        batch_size_kg = scenarios[scenario_id]["profile"]["batch_size"] / 1000.0
        pei_wt = scenarios[scenario_id]["profile"]["pei_composition"]

        for sorbent_type, stages_map in sorbents_map.items():
            # Skip excluded sorbent types in MCS outputs
            if _is_excluded_mcs_sorbent(sorbent_type):
                continue
            
            # Track Stage_Total matrices for grand total calculation (only for Direct)
            stage_total_matrices_dict = {}  # {stage: {category: matrix}}
            
            # Process all stages (Aziridine, PEI, Support, Sorbent) and Total
            items_to_process = [("Total", "Total")]
            # Add only the 4 individual production stages (exclude top-level summaries).
            for stage_key in stages_map.keys():
                if stage_key not in {"Total", "Stage_Total"} and isinstance(stages_map[stage_key], dict):
                    stage_block = stages_map[stage_key]
                    # Process Stage_Total
                    items_to_process.append((stage_key, "Stage_Total"))
                    # Process all individual components
                    for component_key in stage_block.keys():
                        if component_key != "Stage_Total" and isinstance(stage_block[component_key], dict):
                            items_to_process.append((stage_key, component_key))
            
            # Process each stage/component combination
            for current_stage, component_label in items_to_process:
                if current_stage == "Total":
                    impact_block = stages_map.get("Total")
                else:
                    stage_block = stages_map.get(current_stage, {})
                    if isinstance(stage_block, dict):
                        impact_block = stage_block.get(component_label) if component_label in stage_block else None
                    else:
                        impact_block = None

                if not isinstance(impact_block, dict):
                    continue

                for category, matrix in impact_block.items():
                    # Store Stage_Total matrices for grand total calculation
                    if component_label == "Stage_Total" and current_stage != "Total":
                        if current_stage not in stage_total_matrices_dict:
                            stage_total_matrices_dict[current_stage] = {}
                        stage_total_matrices_dict[current_stage][category] = matrix
                    if not (isinstance(matrix, np.ndarray) and matrix.ndim == 3):
                        continue

                    mcs_values = matrix[0, 0, :]
                    n = len(mcs_values)
                    if n == 0:
                        continue

                    ddof = 1 if n > 1 else 0
                    mean_val = float(np.mean(mcs_values))
                    std_val = float(np.std(mcs_values, ddof=ddof))
                    sem_val = std_val / np.sqrt(n) if n > 0 else np.nan
                    sem_percent = (sem_val / mean_val) * 100 if mean_val != 0 else np.nan
                    ci_lower = float(np.percentile(mcs_values, 2.5))
                    ci_upper = float(np.percentile(mcs_values, 97.5))

                    # Calculate per kg sorbent statistics
                    with np.errstate(divide='ignore', invalid='ignore'):
                        per_kg_sorbent_values = np.divide(mcs_values, batch_size_kg, out=np.full_like(mcs_values, np.nan), where=batch_size_kg > 0)
                    per_kg_sorbent_lower = float(np.percentile(per_kg_sorbent_values, 2.5))
                    per_kg_sorbent_mean = float(np.mean(per_kg_sorbent_values))
                    per_kg_sorbent_upper = float(np.percentile(per_kg_sorbent_values, 97.5))

                    pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                    loading = pei_to_loading_map.get(pei_wt, pei_wt)

                    for condition in ["Dry", "Humid"]:
                        per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                        if sorbent_type in adsorption_performance_dictionary:
                            sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                            if (
                                isinstance(sorbent_perf, dict)
                                and "loadings" in sorbent_perf
                                and loading in sorbent_perf["loadings"]
                                and condition in sorbent_perf["loadings"][loading]
                            ):
                                perf = np.asarray(sorbent_perf["loadings"][loading][condition], dtype=float)

                                if perf.shape[0] == n:
                                    # Per kg CO2 captured = per kg sorbent / performance (kg CO2/kg sorbent)
                                    with np.errstate(divide='ignore', invalid='ignore'):
                                        per_capt = np.divide(per_kg_sorbent_values, perf, out=np.full_like(per_kg_sorbent_values, np.nan), where=perf > 0)
                                    per_kg_captured_lower = float(np.percentile(per_capt, 2.5))
                                    per_kg_captured_mean = float(np.mean(per_capt))
                                    per_kg_captured_upper = float(np.percentile(per_capt, 97.5))

                        summary_list.append({
                        "Scenario": scenario_id,
                        "Sorbent Type": sorbent_type,
                        "Stage": current_stage,
                        "Component": component_label,
                        "Impact Category": category,
                        "PEI wt%": pei_wt,
                        "Batch Size (kg)": batch_size_kg,
                        "Condition": condition,
                        "95% CI Lower": ci_lower,
                        "Mean": mean_val,
                        "95% CI Upper": ci_upper,
                        "95% CI Lower (per kg sorbent)": per_kg_sorbent_lower,
                        "Mean (per kg sorbent)": per_kg_sorbent_mean,
                        "95% CI Upper (per kg sorbent)": per_kg_sorbent_upper,
                        "95% CI Lower (per kg CO2 captured)": per_kg_captured_lower,
                        "Mean (per kg CO2 captured)": per_kg_captured_mean,
                        "95% CI Upper (per kg CO2 captured)": per_kg_captured_upper,
                        "Sample Size": n,
                    })
            
            # Calculate grand total across all stages for material impacts
            if stage_total_matrices_dict:
                # Get all impact categories from any stage
                all_categories = set()
                for stage_dict in stage_total_matrices_dict.values():
                    all_categories.update(stage_dict.keys())
                
                # Calculate grand total for each impact category
                for category in all_categories:
                    grand_total_matrix = None
                    reference_shape = None
                    
                    # Get reference shape from first available matrix
                    for stage, category_dict in stage_total_matrices_dict.items():
                        if category in category_dict:
                            matrix = category_dict[category]
                            if matrix is not None:
                                reference_shape = matrix.shape
                                grand_total_matrix = np.zeros(reference_shape)
                                break
                    
                    # Sum all Stage_Total matrices for this category across all stages
                    if grand_total_matrix is not None:
                        for stage, category_dict in stage_total_matrices_dict.items():
                            if category in category_dict:
                                matrix = category_dict[category]
                                if matrix is not None and matrix.shape == reference_shape:
                                    grand_total_matrix += matrix
                                elif matrix is not None:
                                    grand_total_matrix += np.broadcast_to(matrix, reference_shape)
                        
                        # Calculate statistics for grand total
                        mcs_values = grand_total_matrix[0, 0, :]
                        n = len(mcs_values)
                        if n > 0:
                            ddof = 1 if n > 1 else 0
                            mean_val = float(np.mean(mcs_values))
                            ci_lower = float(np.percentile(mcs_values, 2.5))
                            ci_upper = float(np.percentile(mcs_values, 97.5))
                            
                            # Calculate per kg sorbent statistics
                            with np.errstate(divide='ignore', invalid='ignore'):
                                per_kg_sorbent_values = np.divide(mcs_values, batch_size_kg, out=np.full_like(mcs_values, np.nan), where=batch_size_kg > 0)
                            per_kg_sorbent_lower = float(np.percentile(per_kg_sorbent_values, 2.5))
                            per_kg_sorbent_mean = float(np.mean(per_kg_sorbent_values))
                            per_kg_sorbent_upper = float(np.percentile(per_kg_sorbent_values, 97.5))
                            
                            pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                            loading = pei_to_loading_map.get(pei_wt, pei_wt)
                            
                            for condition in ["Dry", "Humid"]:
                                per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                                if sorbent_type in adsorption_performance_dictionary:
                                    sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                                    if (
                                        isinstance(sorbent_perf, dict)
                                        and "loadings" in sorbent_perf
                                        and loading in sorbent_perf["loadings"]
                                        and condition in sorbent_perf["loadings"][loading]
                                    ):
                                        perf = np.asarray(sorbent_perf["loadings"][loading][condition], dtype=float)
                                        if perf.shape[0] == n:
                                            # Per kg CO2 captured = per kg sorbent / performance (kg CO2/kg sorbent)
                                            with np.errstate(divide='ignore', invalid='ignore'):
                                                per_capt = np.divide(per_kg_sorbent_values, perf, out=np.full_like(per_kg_sorbent_values, np.nan), where=perf > 0)
                                            per_kg_captured_lower = float(np.percentile(per_capt, 2.5))
                                            per_kg_captured_mean = float(np.mean(per_capt))
                                            per_kg_captured_upper = float(np.percentile(per_capt, 97.5))
                            
                                summary_list.append({
                                    "Scenario": scenario_id,
                                    "Sorbent Type": sorbent_type,
                                    "Stage": "Total",
                                    "Component": "Total",
                                    "Impact Category": category,
                                    "PEI wt%": pei_wt,
                                    "Batch Size (kg)": batch_size_kg,
                                    "Condition": condition,
                                    "95% CI Lower": ci_lower,
                                    "Mean": mean_val,
                                    "95% CI Upper": ci_upper,
                                    "95% CI Lower (per kg sorbent)": per_kg_sorbent_lower,
                                    "Mean (per kg sorbent)": per_kg_sorbent_mean,
                                    "95% CI Upper (per kg sorbent)": per_kg_sorbent_upper,
                        "95% CI Lower (per kg CO2 captured)": per_kg_captured_lower,
                        "Mean (per kg CO2 captured)": per_kg_captured_mean,
                        "95% CI Upper (per kg CO2 captured)": per_kg_captured_upper,
                        "Sample Size": n,
                    })

    def _safe_sheet_name(name):
        safe = name.replace(" ", "_")
        return safe[:31]
    df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))
    # Remove any rows where Stage is a Stage_Total summary label.
    if "Stage" in df.columns:
        stage_norm = df["Stage"].astype(str).str.strip().str.lower()
        df = df[~stage_norm.str.contains(r"stage[_\\s]*tota", regex=True, na=False)].reset_index(drop=True)
    # Enforce one row per logical key (prevents duplicated totals/components in saved sheet).
    if not df.empty:
        dedup_keys = [
            "Scenario", "Sorbent Type", "Stage", "Component",
            "Impact Category", "PEI wt%", "Batch Size (kg)", "Condition"
        ]
        dedup_keys = [k for k in dedup_keys if k in df.columns]
        if dedup_keys:
            before = len(df)
            df = df.drop_duplicates(subset=dedup_keys, keep="first").reset_index(drop=True)
            removed = before - len(df)
            if removed > 0:
                print(f"ℹ️ Removed {removed} duplicate material-impact rows before saving.")
    sheet_name = _safe_sheet_name(f"LC_Chemical_Impacts")
    try:
        with pd.ExcelWriter("Total_Environmental_Impacts.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Saved {len(df)} rows to sheet '{sheet_name}' in Total_Environmental_Impacts.xlsx")
    except FileNotFoundError:
        with pd.ExcelWriter("Total_Environmental_Impacts.xlsx", engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Created file and saved {len(df)} rows to sheet '{sheet_name}' in Total_Environmental_Impacts.xlsx")

    return df


# ========================================================================================================================
# STEP 3B: Monte Carlo Simulations - Combined Total Impacts Summary (Energy + Materials)
# ========================================================================================================================

def monte_carlo_statistics_total_impacts(
    energy_impact_dictionary,
    material_impact_dictionary,
    adsorption_performance_dictionary,
    scenarios,
):
    """
    Compute MC statistics for combined total impacts.

    Follows the same grand-total logic as material impacts:
    1) Sum material Stage_Total matrices across the 4 production stages.
    2) Add energy electricity total + heat total matrices.
    3) Run MCS statistics on the resulting combined matrix.

    Output is one row per:
      Scenario, Sorbent Type, Impact Category, Condition
    with Stage fixed as "Total".
    """
    summary_list = []
    impact_categories = [
        "Climate Change (kg CO2-Eq)",
        "Fossil resource scarcity (kg oil-Eq)",
        "Water Use (m3)",
    ]
    def _norm_cat(name):
        return str(name).strip().lower()

    def _get_category_matrix(block, category):
        """Get category matrix using exact key, then normalized-key fallback."""
        if not isinstance(block, dict):
            return None
        if category in block:
            return block.get(category)
        target = _norm_cat(category)
        for k, v in block.items():
            if _norm_cat(k) == target:
                return v
        return None

    def _sum_into(acc_matrix, add_matrix):
        """Shape-safe matrix summation for (1,1,mcs) style arrays."""
        if add_matrix is None:
            return acc_matrix
        if not (isinstance(add_matrix, np.ndarray) and add_matrix.ndim == 3):
            return acc_matrix
        if acc_matrix is None:
            return add_matrix.copy()
        if add_matrix.shape == acc_matrix.shape:
            return acc_matrix + add_matrix
        return acc_matrix + np.broadcast_to(add_matrix, acc_matrix.shape)

    def _sum_stage_totals(block, category):
        """
        Fallback helper: sum stage-level Stage_Total matrices for a category.
        Works for both material and energy blocks shaped as
        {stage: {"Stage_Total": {category: matrix}}, "Total": {...}}.
        """
        if not isinstance(block, dict):
            return None
        acc = None
        for stage_name, stage_data in block.items():
            if stage_name == "Total" or not isinstance(stage_data, dict):
                continue
            mat = _get_category_matrix(stage_data.get("Stage_Total", {}), category)
            acc = _sum_into(acc, mat)
        return acc

    for scenario_id, sorbents_map in material_impact_dictionary.items():
        if scenario_id not in scenarios:
            continue
        batch_size_kg = scenarios[scenario_id]["profile"]["batch_size"] / 1000.0
        pei_wt = scenarios[scenario_id]["profile"]["pei_composition"]

        for sorbent_type, stages_map in sorbents_map.items():
            if _is_excluded_mcs_sorbent(sorbent_type):
                continue

            energy_sorbent_block = energy_impact_dictionary.get(scenario_id, {}).get(sorbent_type, {})
            if not isinstance(energy_sorbent_block, dict):
                continue
            us_grid_block = energy_sorbent_block.get("US_Grid_Average", {})
            heat_block = energy_sorbent_block.get("Heat", {})
            if not isinstance(us_grid_block, dict) or not isinstance(heat_block, dict):
                continue
            us_grid_total_block = us_grid_block.get("Total", {})
            heat_total_block = heat_block.get("Total", {})

            pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
            loading = pei_to_loading_map.get(pei_wt, pei_wt)

            # Chemical impacts source of truth requested by user:
            # use material "Total" block per sorbent/scenario.
            material_total_block = stages_map.get("Total", {}) if isinstance(stages_map, dict) else {}
            if not isinstance(material_total_block, dict):
                material_total_block = {}

            all_categories = set(material_total_block.keys()) | set(impact_categories)
            # Include categories present only in energy totals too.
            if isinstance(us_grid_total_block, dict):
                all_categories.update(us_grid_total_block.keys())
            if isinstance(heat_total_block, dict):
                all_categories.update(heat_total_block.keys())

            for category in sorted(all_categories):
                chem_total_matrix = _get_category_matrix(material_total_block, category)
                if chem_total_matrix is None:
                    chem_total_matrix = _sum_stage_totals(stages_map, category)

                elec_total_matrix = _get_category_matrix(us_grid_total_block, category)
                if elec_total_matrix is None:
                    elec_total_matrix = _sum_stage_totals(us_grid_block, category)

                heat_total_matrix = _get_category_matrix(heat_total_block, category)
                if heat_total_matrix is None:
                    heat_total_matrix = _sum_stage_totals(heat_block, category)

                # IMPORTANT:
                # - Material impact matrices are absolute totals -> normalize by batch size.
                # - Energy impact matrices are already normalized per kg sorbent.
                chem_perkg_matrix = None
                if isinstance(chem_total_matrix, np.ndarray):
                    if batch_size_kg > 0:
                        chem_perkg_matrix = chem_total_matrix / batch_size_kg

                total_matrix = None
                total_matrix = _sum_into(total_matrix, chem_perkg_matrix)
                total_matrix = _sum_into(total_matrix, elec_total_matrix)
                total_matrix = _sum_into(total_matrix, heat_total_matrix)
                if total_matrix is None:
                    continue

                per_kg_sorbent_values = total_matrix[0, 0, :]
                n = len(per_kg_sorbent_values)
                if n == 0:
                    continue

                ci_lower = float(np.percentile(per_kg_sorbent_values, 2.5))
                per_kg_sorbent_lower = float(np.percentile(per_kg_sorbent_values, 2.5))
                per_kg_sorbent_mean = float(np.mean(per_kg_sorbent_values))
                per_kg_sorbent_upper = float(np.percentile(per_kg_sorbent_values, 97.5))

                for condition in ["Dry", "Humid"]:
                    per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                    if sorbent_type in adsorption_performance_dictionary:
                        sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                        if (                                isinstance(sorbent_perf, dict)
                                and "loadings" in sorbent_perf
                                and loading in sorbent_perf["loadings"]
                                and condition in sorbent_perf["loadings"][loading]
                            ):
                                perf = np.asarray(sorbent_perf["loadings"][loading][condition], dtype=float)
                                if perf.shape[0] == n:
                                    # Per kg CO2 captured = per kg sorbent / performance (kg CO2/kg sorbent)
                                    with np.errstate(divide='ignore', invalid='ignore'):
                                        per_capt = np.divide(per_kg_sorbent_values, perf, out=np.full_like(per_kg_sorbent_values, np.nan), where=perf > 0)
                                    per_kg_captured_lower = float(np.percentile(per_capt, 2.5))
                                    per_kg_captured_mean = float(np.mean(per_capt))
                                    per_kg_captured_upper = float(np.percentile(per_capt, 97.5))


                    summary_list.append({
                        "Scenario": scenario_id,
                        "Sorbent Type": sorbent_type,
                        "Stage": "Total",
                        "Impact Category": category,
                        "PEI wt%": pei_wt,
                        "Batch Size (kg)": batch_size_kg,
                        "Condition": condition,
                        "95% CI Lower (per kg sorbent)": per_kg_sorbent_lower,
                        "Mean (per kg sorbent)": per_kg_sorbent_mean,
                        "95% CI Upper (per kg sorbent)": per_kg_sorbent_upper,
                        "95% CI Lower (per kg CO2 captured)": per_kg_captured_lower,
                        "Mean (per kg CO2 captured)": per_kg_captured_mean,
                        "95% CI Upper (per kg CO2 captured)": per_kg_captured_upper,
                    })

    def _safe_sheet_name(name):
        safe = name.replace(" ", "_")
        return safe[:31]

    df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))
    if not df.empty:
        dedup_keys = [
            "Scenario", "Sorbent Type", "Stage", "Impact Category",
            "PEI wt%", "Batch Size (kg)", "Condition"
        ]
        dedup_keys = [k for k in dedup_keys if k in df.columns]
        if dedup_keys:
            df = df.drop_duplicates(subset=dedup_keys, keep="last").reset_index(drop=True)

    output_file = "Total_Environmental_Impacts.xlsx"
    climate_sheet = _safe_sheet_name("Climate Change - Summary")
    fossil_sheet = _safe_sheet_name("Fossil resource scarcity - Summary")
    water_sheet = _safe_sheet_name("Water Use - Summary")

    # Remove legacy LC_Total_Impacts* sheets so only the three summary sheets remain for totals.
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        legacy_sheets = [s for s in wb.sheetnames if str(s).startswith("LC_Total_Impacts")]
        for s in legacy_sheets:
            del wb[s]
        if legacy_sheets:
            wb.save(output_file)
            print(f"ℹ️ Removed legacy sheets: {legacy_sheets}")
    except FileNotFoundError:
        pass

    climate_df = df.iloc[0:0].copy()
    fossil_df = df.iloc[0:0].copy()
    water_df = df.iloc[0:0].copy()
    if not df.empty and "Impact Category" in df.columns:
        climate_df = df[df["Impact Category"] == "Climate Change (kg CO2-Eq)"].reset_index(drop=True)
        fossil_df = df[df["Impact Category"] == "Fossil resource scarcity (kg oil-Eq)"].reset_index(drop=True)
        water_df = df[df["Impact Category"] == "Water Use (m3)"].reset_index(drop=True)

    def _sort_for_summary(out_df):
        if out_df.empty:
            return out_df
        sort_priority = [
            "Sorbent Type",
            "PEI wt%",
            "Scenario",
            "Batch Size (kg)",
            "Condition",
            "Impact Category",
        ]
        sort_cols = [c for c in sort_priority if c in out_df.columns]
        if not sort_cols:
            return out_df
        return out_df.sort_values(sort_cols, kind="stable").reset_index(drop=True)

    climate_df = _sort_for_summary(climate_df)
    fossil_df = _sort_for_summary(fossil_df)
    water_df = _sort_for_summary(water_df)

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            climate_df.to_excel(writer, sheet_name=climate_sheet, index=False)
            fossil_df.to_excel(writer, sheet_name=fossil_sheet, index=False)
            water_df.to_excel(writer, sheet_name=water_sheet, index=False)
            print(f"✅ Saved {len(climate_df)} rows to sheet '{climate_sheet}' in {output_file}")
            print(f"✅ Saved {len(fossil_df)} rows to sheet '{fossil_sheet}' in {output_file}")
            print(f"✅ Saved {len(water_df)} rows to sheet '{water_sheet}' in {output_file}")
    except FileNotFoundError:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            climate_df.to_excel(writer, sheet_name=climate_sheet, index=False)
            fossil_df.to_excel(writer, sheet_name=fossil_sheet, index=False)
            water_df.to_excel(writer, sheet_name=water_sheet, index=False)
            print(f"✅ Created file and saved {len(climate_df)} rows to sheet '{climate_sheet}' in {output_file}")
            print(f"✅ Created file and saved {len(fossil_df)} rows to sheet '{fossil_sheet}' in {output_file}")
            print(f"✅ Created file and saved {len(water_df)} rows to sheet '{water_sheet}' in {output_file}")

    return df

# ========================================================================================================================
# STEP 4: Monte Carlo Simulations - Direct Energy Demand
# ========================================================================================================================

def monte_carlo_statistics_direct_energy(energy_dictionary, adsorption_performance_dictionary, stage_name, scenarios, unit_label="kg sorbent", pei_mass_dic=None, support_mass_dic=None):
    """
    Calculate Monte Carlo statistics for direct energy demand.
    
    Parameters:
    -----------
    direct_energy_dictionary : dict
        Nested dictionary with structure:
        {scenario_id: {sorbent_type: {energy_type: {stage: {process: 3D_matrix}}}}}
        where 3D_matrix has shape (1, 1, mcs_number) for direct energy data
    stage_name : str
        Name for the stage (e.g., "Aziridine", "PEI", "Support", "Sorbent")
    unit_label : str
        Label for normalization (default: "kg sorbent"). Use "kg product" for per-kg-product dictionaries.
    scenarios : dict
        Scenarios dictionary containing batch_size and pei_composition information
    pei_mass_dic : dict, optional
        PEI mass dictionary to get PEI produced (for per kg CO2 calculation when unit is product)
    support_mass_dic : dict, optional
        Support mass dictionary to get Support produced (for per kg CO2 calculation when unit is product)
        
    Returns:
    --------
    pd.DataFrame
        DataFrame with statistics for each scenario/sorbent/process combination
    """
    
    def _safe_sheet_name(name):
        safe = name.replace(" ", "_")
        return safe[:31]
    
    def _get_pei_produced(scenario, sorbent, mcs_len):
        """Get PEI produced at step 30 (kg)"""
        pei_produced = np.full(mcs_len, np.nan, dtype=float)
        if pei_mass_dic and scenario in pei_mass_dic and sorbent in pei_mass_dic[scenario]:
            pei_masses = pei_mass_dic[scenario][sorbent]
            if "Polyethyleneimine" in pei_masses:
                arr = np.asarray(pei_masses["Polyethyleneimine"], dtype=float)
                if arr.ndim == 3 and arr.shape[1] > 30:
                    pei_produced = arr[0, 30, :mcs_len] / 1000.0  # g -> kg
        return pei_produced
    
    def _get_support_produced(scenario, sorbent, mcs_len):
        """Get Support produced (kg) - last non-zero value"""
        support_produced = np.full(mcs_len, np.nan, dtype=float)
        if support_mass_dic and scenario in support_mass_dic and sorbent in support_mass_dic[scenario]:
            support_masses = support_mass_dic[scenario][sorbent]
            sorbent_lower = sorbent.lower()
            candidates = []
            if "silica" in sorbent_lower:
                candidates = [c for c in support_masses.keys() if "silica" in c.lower()]
            elif "polyhipe" in sorbent_lower:
                candidates = [c for c in support_masses.keys() if "polyhipe" in c.lower()]
            elif "mil" in sorbent_lower:
                candidates = ["MIL-101(Cr)"] if "MIL-101(Cr)" in support_masses else [
                    c for c in support_masses.keys() if "mil-101(cr)" in c.lower()
                ]
            elif "alumina" in sorbent_lower:
                candidates = [c for c in support_masses.keys() if "alumina" in c.lower()]
            if candidates:
                comp = candidates[0]
                arr = np.asarray(support_masses[comp], dtype=float)
                if arr.ndim == 3 and arr.shape[2] >= mcs_len:
                    for k in range(mcs_len):
                        col = arr[0, :, k]
                        nz = np.flatnonzero(col > 0)
                        if nz.size > 0:
                            support_produced[k] = col[nz[-1]] / 1000.0  # g -> kg
        return support_produced

    def _is_recovery_process(process_name):
        """Match solvent/byproduct recovery process labels."""
        p = str(process_name).strip().lower()
        return (
            ("recovery" in p and "solvent" in p)
            or ("recovery" in p and "byproduct" in p)
        )

    summary_list = []
    # For separate Excel with total direct energy per batch
    batch_total_summary_list = []
    unit_label_norm = str(unit_label).strip().lower()
    unit_type = "sorbent" if "sorbent" in unit_label_norm else ("product" if "product" in unit_label_norm else "other")
    include_all_stages = stage_name in ["Total", "Direct", "All"]

    def _build_energy_summary_rows(matrix, scenario_id, sorbent_type, stage, pei_wt, batch_size_kg, energy_type_label, process_label):
        """Build one or more summary rows (Dry only for sorbent-normalized output)."""
        if not (isinstance(matrix, np.ndarray) and matrix.ndim == 3):
            return []
        mcs_values = matrix[0, 0, :]
        num_samples = len(mcs_values)
        if num_samples == 0:
            return []

        mean_val = np.mean(mcs_values)
        ci_lower = np.percentile(mcs_values, 2.5)
        ci_upper = np.percentile(mcs_values, 97.5)

        pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
        loading = pei_to_loading_map.get(pei_wt, pei_wt)
        conditions_to_process = ["Dry"] if unit_type == "sorbent" else [None]

        rows = []
        for condition in conditions_to_process:
            per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
            if unit_type == "sorbent" and sorbent_type in adsorption_performance_dictionary:
                sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                if (
                    isinstance(sorbent_perf, dict)
                    and "loadings" in sorbent_perf
                    and loading in sorbent_perf["loadings"]
                    and condition in sorbent_perf["loadings"][loading]
                ):
                    performance_arrays = sorbent_perf["loadings"][loading][condition]
                    if performance_arrays is not None and performance_arrays.shape[0] == num_samples:
                        energy_per_kg_captured_values = mcs_values / performance_arrays
                        per_kg_captured_lower = np.percentile(energy_per_kg_captured_values, 2.5)
                        per_kg_captured_mean = np.mean(energy_per_kg_captured_values)
                        per_kg_captured_upper = np.percentile(energy_per_kg_captured_values, 97.5)

            summary_dict = {
                'Stage': stage,
                'Scenario': scenario_id,
                'Sorbent Type': sorbent_type,
                'Energy Type': energy_type_label,
                'Process': process_label,
                'PEI wt%': pei_wt,
                'Batch Size (kg)': batch_size_kg,
                f'95% CI Lower (kWh or MJ/{unit_label})': ci_lower,
                f'Mean (kWh or MJ/{unit_label})': mean_val,
                f'95% CI Upper (kWh or MJ/{unit_label})': ci_upper,
                'Sample Size': num_samples,
            }
            if unit_type == "sorbent":
                summary_dict['Condition'] = condition
                summary_dict['95% CI Lower (kWh or MJ/kg CO2 captured)'] = per_kg_captured_lower
                summary_dict['Mean (kWh or MJ/kg CO2 captured)'] = per_kg_captured_mean
                summary_dict['95% CI Upper (kWh or MJ/kg CO2 captured)'] = per_kg_captured_upper
            rows.append(summary_dict)
        return rows
    
    # Iterate through scenarios
    for scenario_id, scenario_data in energy_dictionary.items():
        # Get scenario information
        batch_size_kg = scenarios[scenario_id]['profile']['batch_size']/1000  # Convert to kg
        pei_wt = scenarios[scenario_id]['profile']['pei_composition']
        
        # Determine whether to only calculate totals (for scenarios > 20)
        try:
            scenario_num = int(scenario_id)
        except (TypeError, ValueError):
            scenario_num = None
        totals_only = scenario_num is not None and scenario_num > 20
        # Lab-scale assumption: scenarios 1-5 have no solvent/byproduct recovery.
        exclude_recovery_processes = scenario_num is not None and 1 <= scenario_num <= 5

        # Iterate through sorbent types
        for sorbent_type, sorbent_data in scenario_data.items():
            # Skip excluded sorbent types in MCS outputs
            if _is_excluded_mcs_sorbent(sorbent_type):
                continue
            stage_totals = {}  # Will store {stage: {"electricity": [matrices], "heat": [matrices]}}
            stage_total_matrices_dict = {}  # Will store {stage: Stage_Total_matrix} for grand total calculation
            # Iterate through energy types (electricity/heat) - primary structure
            for energy_type, stages_map in sorbent_data.items():
                if energy_type not in ["electricity", "heat"]:
                    continue
                if not isinstance(stages_map, dict):
                    continue

                for stage, processes in stages_map.items():
                    if not include_all_stages and stage != stage_name:
                        continue
                    if not isinstance(processes, dict):
                        continue

                    for process, matrix in processes.items():
                        # Skip Stage_Total - it's already calculated in the dictionary
                        if process == "Stage_Total":
                            continue
                        if exclude_recovery_processes and _is_recovery_process(process):
                            continue
                        if not (isinstance(matrix, np.ndarray) and matrix.ndim == 3):
                            continue

                        mcs_values = matrix[0, 0, :]
                        num_samples = len(mcs_values)
                        if num_samples == 0:
                            continue

                        # accumulate stage totals by energy type (for recalculating Stage_Total if needed)
                        if stage not in stage_totals:
                            stage_totals[stage] = {"electricity": [], "heat": []}
                        stage_totals[stage][energy_type].append(matrix)

                        # Calculate statistics (per unit: kg sorbent or kg product)
                        mean_val = np.mean(mcs_values)
                        std_val = np.std(mcs_values)
                        sem_val = std_val / np.sqrt(num_samples)
                        sem_percent = (sem_val / mean_val) * 100 if mean_val != 0 else np.nan
                        ci_lower = np.percentile(mcs_values, 2.5)
                        ci_upper = np.percentile(mcs_values, 97.5)

                        # Calculate per kg captured statistics only for sorbent normalization
                        pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                        loading = pei_to_loading_map.get(pei_wt, pei_wt)

                        # Keep one energy row per process/stage for sorbent-normalized outputs.
                        # (Dry/Humid affects only adsorption-based denominator, not direct energy itself.)
                        conditions_to_process = ["Dry"] if unit_type == "sorbent" else [None]
                        
                        # For scenarios <= 20, report by-process statistics; for scenarios > 20, skip per-process rows
                        if not totals_only:
                            for condition in conditions_to_process:
                                per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                                # Only calculate per kg CO2 captured for sorbent normalization
                                if unit_type == "sorbent" and sorbent_type in adsorption_performance_dictionary:
                                    sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                                    if (
                                        isinstance(sorbent_perf, dict)
                                        and "loadings" in sorbent_perf
                                        and loading in sorbent_perf["loadings"]
                                        and condition in sorbent_perf["loadings"][loading]
                                    ):
                                        perf_values = sorbent_perf["loadings"][loading][condition]
                                        performance_arrays = perf_values
                                        if performance_arrays is not None and performance_arrays.shape[0] == num_samples:
                                            # For per kg sorbent normalization, use performance directly
                                            energy_per_kg_captured_values = mcs_values / performance_arrays
                                            per_kg_captured_lower = np.percentile(energy_per_kg_captured_values, 2.5)
                                            per_kg_captured_mean = np.mean(energy_per_kg_captured_values)
                                            per_kg_captured_upper = np.percentile(energy_per_kg_captured_values, 97.5)

                                # Build summary dictionary
                                summary_dict = {
                                    'Stage': stage,
                                    'Scenario': scenario_id,
                                    'Sorbent Type': sorbent_type,
                                    'Energy Type': energy_type,
                                    'Process': process,
                                    'PEI wt%': pei_wt,
                                    'Batch Size (kg)': batch_size_kg,
                                    f'95% CI Lower (kWh or MJ/{unit_label})': ci_lower,
                                    f'Mean (kWh or MJ/{unit_label})': mean_val,
                                    f'95% CI Upper (kWh or MJ/{unit_label})': ci_upper,
                                    'Sample Size': num_samples,
                                }
                                # Only add Condition column for sorbent normalization
                                if unit_type == "sorbent":
                                    summary_dict['Condition'] = condition
                                # Only add per kg CO2 captured columns for sorbent normalization
                                if unit_type == "sorbent":
                                    summary_dict['95% CI Lower (kWh or MJ/kg CO2 captured)'] = per_kg_captured_lower
                                    summary_dict['Mean (kWh or MJ/kg CO2 captured)'] = per_kg_captured_mean
                                    summary_dict['95% CI Upper (kWh or MJ/kg CO2 captured)'] = per_kg_captured_upper
                                
                                summary_list.append(summary_dict)

            # Process Stage_Total from dictionary (if exists) or calculate from individual processes
            # Get stages to process: all stages if include_all_stages, otherwise just the requested stage
            if include_all_stages:
                # Get all stages from the dictionary
                stages_to_process = set()
                if "electricity" in sorbent_data:
                    stages_to_process.update(sorbent_data["electricity"].keys())
                if "heat" in sorbent_data:
                    stages_to_process.update(sorbent_data["heat"].keys())
            else:
                # Process only the requested stage
                stages_to_process = {stage_name} if stage_name else set()
            
            for stage in stages_to_process:
                # Stage-level split totals for direct-energy report:
                # "Total Electricity" -> "Electricity", "Total Heat" -> "Heat"
                if not totals_only and stage in stage_totals:
                    energy_matrices = stage_totals[stage]
                    electricity_matrices = energy_matrices.get("electricity", [])
                    heat_matrices = energy_matrices.get("heat", [])

                    reference_shape = None
                    if electricity_matrices:
                        reference_shape = electricity_matrices[0].shape
                    elif heat_matrices:
                        reference_shape = heat_matrices[0].shape

                    if reference_shape is not None:
                        if electricity_matrices:
                            total_electricity_matrix = np.zeros(reference_shape)
                            for matrix in electricity_matrices:
                                if matrix.shape == reference_shape:
                                    total_electricity_matrix += matrix
                                else:
                                    total_electricity_matrix += np.broadcast_to(matrix, reference_shape)
                            summary_list.extend(
                                _build_energy_summary_rows(
                                    total_electricity_matrix,
                                    scenario_id,
                                    sorbent_type,
                                    stage,
                                    pei_wt,
                                    batch_size_kg,
                                    "Total Electricity",
                                    "Electricity",
                                )
                            )

                        if heat_matrices:
                            total_heat_matrix = np.zeros(reference_shape)
                            for matrix in heat_matrices:
                                if matrix.shape == reference_shape:
                                    total_heat_matrix += matrix
                                else:
                                    total_heat_matrix += np.broadcast_to(matrix, reference_shape)
                            summary_list.extend(
                                _build_energy_summary_rows(
                                    total_heat_matrix,
                                    scenario_id,
                                    sorbent_type,
                                    stage,
                                    pei_wt,
                                    batch_size_kg,
                                    "Total Heat",
                                    "Heat",
                                )
                            )

                # Try to get Stage_Total directly from dictionary (already in MJ)
                stage_total_matrix = None
                # For scenarios 1-5, force recomputation from filtered per-process rows so
                # recovery does not remain embedded in precomputed Stage_Total matrices.
                if not exclude_recovery_processes:
                    if "electricity" in sorbent_data and stage in sorbent_data["electricity"]:
                        if "Stage_Total" in sorbent_data["electricity"][stage]:
                            stage_total_matrix = sorbent_data["electricity"][stage]["Stage_Total"]
                    if stage_total_matrix is None and "heat" in sorbent_data and stage in sorbent_data["heat"]:
                        if "Stage_Total" in sorbent_data["heat"][stage]:
                            stage_total_matrix = sorbent_data["heat"][stage]["Stage_Total"]
                
                # If Stage_Total not in dictionary, calculate from individual processes
                if stage_total_matrix is None and stage in stage_totals:
                    energy_matrices = stage_totals[stage]
                    # Get electricity and heat matrices separately
                    electricity_matrices = energy_matrices.get("electricity", [])
                    heat_matrices = energy_matrices.get("heat", [])
                    
                    if electricity_matrices or heat_matrices:
                        # Get reference shape from first available matrix
                        reference_shape = None
                        if electricity_matrices:
                            reference_shape = electricity_matrices[0].shape
                        elif heat_matrices:
                            reference_shape = heat_matrices[0].shape
                        
                        if reference_shape is not None:
                            # Sum electricity matrices (in kWh) and convert to MJ
                            total_electricity_MJ = np.zeros(reference_shape)
                            if electricity_matrices:
                                for matrix in electricity_matrices:
                                    if matrix.shape == reference_shape:
                                        # Convert electricity from kWh to MJ (1 kWh = 3.6 MJ)
                                        total_electricity_MJ += matrix * 3.6
                                    else:
                                        total_electricity_MJ += np.broadcast_to(matrix * 3.6, reference_shape)
                            
                            # Sum heat matrices (already in MJ)
                            total_heat_MJ = np.zeros(reference_shape)
                            if heat_matrices:
                                for matrix in heat_matrices:
                                    if matrix.shape == reference_shape:
                                        total_heat_MJ += matrix
                                    else:
                                        total_heat_MJ += np.broadcast_to(matrix, reference_shape)
                            
                            # Total energy in MJ (electricity converted + heat)
                            stage_total_matrix = total_electricity_MJ + total_heat_MJ
                
                # Process Stage_Total if we have it (keep in MJ, no conversion)
                if stage_total_matrix is not None:
                    # Store Stage_Total matrix for grand total calculation (only for Direct)
                    if include_all_stages:
                        stage_total_matrices_dict[stage] = stage_total_matrix
                    
                    # Stage_Total is already in MJ in dictionary, keep it in MJ
                    mcs_values = stage_total_matrix[0, 0, :]
                    num_samples = len(mcs_values)
                    if num_samples == 0:
                        continue

                    mean_val = np.mean(mcs_values)
                    std_val = np.std(mcs_values)
                    sem_val = std_val / np.sqrt(num_samples)
                    sem_percent = (sem_val / mean_val) * 100 if mean_val != 0 else np.nan
                    ci_lower = np.percentile(mcs_values, 2.5)
                    ci_upper = np.percentile(mcs_values, 97.5)

                    pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                    loading = pei_to_loading_map.get(pei_wt, pei_wt)

                    # Keep one energy row per process/stage for sorbent-normalized outputs.
                    conditions_to_process = ["Dry"] if unit_type == "sorbent" else [None]
                    
                    # For scenarios <= 20, report Stage_Total per stage; for scenarios > 20, skip per-stage totals
                    if not totals_only:
                        for condition in conditions_to_process:
                            per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                            # Only calculate per kg CO2 captured for sorbent normalization
                            if unit_type == "sorbent" and sorbent_type in adsorption_performance_dictionary:
                                sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                                if (
                                    isinstance(sorbent_perf, dict)
                                    and "loadings" in sorbent_perf
                                    and loading in sorbent_perf["loadings"]
                                    and condition in sorbent_perf["loadings"][loading]
                                ):
                                    perf_values = sorbent_perf["loadings"][loading][condition]
                                    performance_arrays = perf_values
                                    if performance_arrays is not None and performance_arrays.shape[0] == num_samples:
                                        # For per kg sorbent normalization, use performance directly
                                        energy_per_kg_captured_values = mcs_values / performance_arrays
                                        per_kg_captured_lower = np.percentile(energy_per_kg_captured_values, 2.5)
                                        per_kg_captured_mean = np.mean(energy_per_kg_captured_values)
                                        per_kg_captured_upper = np.percentile(energy_per_kg_captured_values, 97.5)

                            # Build summary dictionary
                            summary_dict = {
                                'Stage': stage,
                                'Scenario': scenario_id,
                                'Sorbent Type': sorbent_type,
                                'Energy Type': "Total",
                                'Process': "Stage_Total",
                                'PEI wt%': pei_wt,
                                'Batch Size (kg)': batch_size_kg,
                                f'95% CI Lower (kWh or MJ/{unit_label})': ci_lower,
                                f'Mean (kWh or MJ/{unit_label})': mean_val,
                                f'95% CI Upper (kWh or MJ/{unit_label})': ci_upper,
                                'Sample Size': num_samples,
                            }
                            # Only add Condition column for sorbent normalization
                            if unit_type == "sorbent":
                                summary_dict['Condition'] = condition
                            # Only add per kg CO2 captured columns for sorbent normalization
                            if unit_type == "sorbent":
                                summary_dict['95% CI Lower (kWh or MJ/kg CO2 captured)'] = per_kg_captured_lower
                                summary_dict['Mean (kWh or MJ/kg CO2 captured)'] = per_kg_captured_mean
                                summary_dict['95% CI Upper (kWh or MJ/kg CO2 captured)'] = per_kg_captured_upper
                            
                            summary_list.append(summary_dict)
            
            # Calculate grand total across all stages for Direct energy (only when include_all_stages is True)
            if include_all_stages and stage_total_matrices_dict:
                # Sum all Stage_Total matrices from all stages
                grand_total_matrix = None
                reference_shape = None
                
                # Get reference shape from first stage
                for stage, matrix in stage_total_matrices_dict.items():
                    if matrix is not None:
                        reference_shape = matrix.shape
                        grand_total_matrix = np.zeros(reference_shape)
                        break
                
                # Sum all stage totals
                if grand_total_matrix is not None:
                    for stage, matrix in stage_total_matrices_dict.items():
                        if matrix is not None and matrix.shape == reference_shape:
                            grand_total_matrix += matrix
                        elif matrix is not None:
                            grand_total_matrix += np.broadcast_to(matrix, reference_shape)
                    
                    # Calculate statistics for grand total
                    mcs_values = grand_total_matrix[0, 0, :]
                    num_samples = len(mcs_values)
                    if num_samples > 0:
                        mean_val = np.mean(mcs_values)
                        ci_lower = np.percentile(mcs_values, 2.5)
                        ci_upper = np.percentile(mcs_values, 97.5)

                        # Map PEI wt% to loading for adsorption performance lookup
                        pei_to_loading_map = {20: 20, 30: 30, 40: 40, 50: 50, 60: 60}
                        loading = pei_to_loading_map.get(pei_wt, pei_wt)

                        # Keep one energy row per process/stage for sorbent-normalized outputs.
                        conditions_to_process = ["Dry"] if unit_type == "sorbent" else [None]

                        # 1) Add grand-total per-unit rows to the main Life_Cycle_Inventory output
                        #    Only for scenarios <= 20 (i.e., when not totals_only)
                        if not totals_only:
                            for condition in conditions_to_process:
                                per_kg_captured_lower = per_kg_captured_mean = per_kg_captured_upper = np.nan
                                # Only calculate per kg CO2 captured for sorbent normalization
                                if unit_type == "sorbent" and sorbent_type in adsorption_performance_dictionary:
                                    sorbent_perf = adsorption_performance_dictionary[sorbent_type]
                                    if (
                                        isinstance(sorbent_perf, dict)
                                        and "loadings" in sorbent_perf
                                        and loading in sorbent_perf["loadings"]
                                        and condition in sorbent_perf["loadings"][loading]
                                    ):
                                        perf_values = sorbent_perf["loadings"][loading][condition]
                                        performance_arrays = perf_values
                                        if performance_arrays is not None and performance_arrays.shape[0] == num_samples:
                                            # For per kg sorbent normalization, use performance directly
                                            energy_per_kg_captured_values = mcs_values / performance_arrays
                                            per_kg_captured_lower = np.percentile(energy_per_kg_captured_values, 2.5)
                                            per_kg_captured_mean = np.mean(energy_per_kg_captured_values)
                                            per_kg_captured_upper = np.percentile(energy_per_kg_captured_values, 97.5)

                                # Build summary dictionary for grand total (per-unit energy)
                                summary_dict = {
                                    'Stage': "Total",
                                    'Scenario': scenario_id,
                                    'Sorbent Type': sorbent_type,
                                    'Energy Type': "Total",
                                    'Process': "Total",
                                    'PEI wt%': pei_wt,
                                    'Batch Size (kg)': batch_size_kg,
                                    f'95% CI Lower (kWh or MJ/{unit_label})': ci_lower,
                                    f'Mean (kWh or MJ/{unit_label})': mean_val,
                                    f'95% CI Upper (kWh or MJ/{unit_label})': ci_upper,
                                    'Sample Size': num_samples,
                                }
                                # Only add Condition column for sorbent normalization
                                if unit_type == "sorbent":
                                    summary_dict['Condition'] = condition
                                # Only add per kg CO2 captured columns for sorbent normalization
                                if unit_type == "sorbent":
                                    summary_dict['95% CI Lower (kWh or MJ/kg CO2 captured)'] = per_kg_captured_lower
                                    summary_dict['Mean (kWh or MJ/kg CO2 captured)'] = per_kg_captured_mean
                                    summary_dict['95% CI Upper (kWh or MJ/kg CO2 captured)'] = per_kg_captured_upper

                                summary_list.append(summary_dict)

                        # 2) Add a single grand-total row to the separate batch-size file
                        # (no per kg CO2 captured, no Dry/Humid split)
                        # --- Additional summary: total direct energy per batch size (MJ/{unit_label}) ---
                        # NOTE: This is NOT normalized per kg CO2 captured and does not include Dry/Humid.
                        # Scale by 1000 relative to the per-unit values currently stored in mcs_values.
                        batch_mcs_values = mcs_values 
                        batch_ci_lower = np.percentile(batch_mcs_values, 2.5)
                        batch_mean_val = np.mean(batch_mcs_values)
                        batch_ci_upper = np.percentile(batch_mcs_values, 97.5)

                        batch_summary = {
                            'Scenario': scenario_id,
                            'Sorbent Type': sorbent_type,
                            'PEI wt%': pei_wt,
                            'Batch Size (kg)': batch_size_kg,
                            f'95% CI Lower (MJ/{unit_label})': batch_ci_lower,
                            f'Mean (MJ/{unit_label})': batch_mean_val,
                            f'95% CI Upper (MJ/{unit_label})': batch_ci_upper,
                            'Sample Size': num_samples,
                        }
                        batch_total_summary_list.append(batch_summary)
    
    # Create DataFrame with all statistics
    MC_stats_df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))

    # Create DataFrame for total direct energy per batch (MJ/batch)
    batch_totals_df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(batch_total_summary_list))
    # Sort grand totals by Sorbent Type, then by Batch Size (kg)
    if not batch_totals_df.empty:
        if 'Batch Size (kg)' in batch_totals_df.columns:
            batch_totals_df = batch_totals_df.sort_values(['Sorbent Type', 'Batch Size (kg)']).reset_index(drop=True)
        else:
            batch_totals_df = batch_totals_df.sort_values(['Sorbent Type']).reset_index(drop=True)
    
    # Sort by Scenario, then Sorbent Type, then Stage, then Energy Type (electricity, heat)
    energy_type_order = {
        'electricity': 0,
        'heat': 1,
        'Total Electricity': 2,
        'Total Heat': 3,
        'Total': 4,
    }
    stage_order = {'Aziridine': 0, 'PEI': 1, 'Support': 2, 'Sorbent': 3, 'Total': 4}
    MC_stats_df['_energy_sort'] = MC_stats_df['Energy Type'].map(energy_type_order).fillna(99)
    MC_stats_df['_stage_sort'] = MC_stats_df['Stage'].map(stage_order).fillna(99)
    MC_stats_df = MC_stats_df.sort_values(['Scenario', 'Sorbent Type', '_stage_sort', '_energy_sort']).drop(columns=['_energy_sort', '_stage_sort']).reset_index(drop=True)

    # For Life Cycle Inventory output, exclude scenarios > 20
    MC_stats_LCI_df = MC_stats_df.copy()
    if 'Scenario' in MC_stats_LCI_df.columns:
        def _scenario_leq_20(x):
            try:
                return int(x) <= 20
            except (TypeError, ValueError):
                # Keep non-integer scenario IDs (e.g., labels) by default
                return True
        MC_stats_LCI_df = MC_stats_LCI_df[MC_stats_LCI_df['Scenario'].apply(_scenario_leq_20)].reset_index(drop=True)
    
    # Save full direct energy statistics (filtered to scenarios ≤ 20) to Life_Cycle_Inventory.xlsx
    sheet_suffix = "product" if unit_type == "product" else "sorbent"
    stage_label = "Direct" if stage_name == "Direct" else stage_name
    sheet_name = _safe_sheet_name(f"{stage_label}_Energy_{sheet_suffix}")
    try:
        # Try to append to existing file
        with pd.ExcelWriter('Life_Cycle_Inventory.xlsx', engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            MC_stats_LCI_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Saved {len(MC_stats_LCI_df)} rows to {sheet_name} sheet in Life_Cycle_Inventory.xlsx")
    except FileNotFoundError:
        # Create new file if it doesn't exist
        with pd.ExcelWriter('Life_Cycle_Inventory.xlsx', engine="openpyxl") as writer:
            MC_stats_LCI_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Created new file and saved {len(MC_stats_LCI_df)} rows to {sheet_name} sheet in Life_Cycle_Inventory.xlsx")
    
    return MC_stats_df

# ========================================================================================================================
# STEP 5: Monte Carlo Simulations - Adsorption Performance Results
# ========================================================================================================================

def monte_carlo_statistics_adsorption_performance(performance_results, mcs_number):
    """
    Calculate Monte Carlo statistics for adsorption performance results.
    
    Parameters:
    -----------
    performance_results : dict
        Dictionary containing performance results from adsorption_performance()
        Structure: {sorbent_type: {"N": array, "k": array, "loadings": {loading: {condition: array}}}}
    mcs_number : int
        Number of Monte Carlo samples
        
    Returns:
    --------
    pd.DataFrame
        DataFrame with statistics for each sorbent/loading/condition combination
        Columns: Sorbent_Type, PEI_Loading, Condition, N_low, N_mean, N_upper, K_low, K_mean, K_upper, 
                 Performance_low, Performance_mean, Performance_upper
    """
    
    summary_list = []
    
    # Iterate through sorbent types
    for sorbent_type, sorbent_data in performance_results.items():
        # Skip excluded sorbent types in MCS outputs
        if _is_excluded_mcs_sorbent(sorbent_type):
            continue
        # Get N and k distributions (same for all loadings/conditions for this sorbent)
        N = sorbent_data["N"]
        k = sorbent_data["k"]
        
        # Calculate statistics for N and k once per sorbent
        N_lower = np.percentile(N, 2.5)
        N_mean = np.mean(N)
        N_upper = np.percentile(N, 97.5)

        K_lower = np.percentile(k, 2.5)
        K_mean = np.mean(k)
        K_upper = np.percentile(k, 97.5)
        
        # Extract statistics for each loading and condition
        for loading, conds in sorbent_data["loadings"].items():
            for condition, values in conds.items():
                Performance_lower = np.percentile(values, 2.5)
                Performance_mean = np.mean(values)
                Performance_upper = np.percentile(values, 97.5)
                    
                summary_list.append({
                        'Sorbent_Type': sorbent_type,
                        'PEI_Loading': f"{loading}%",
                        'Condition': condition,
                        '95% CI Lower N (cycles)': N_lower,
                        'Mean N (cycles)': N_mean,
                        '95% CI Upper N (cycles)': N_upper,
                        '95% CI Lower K (1/cycles)': K_lower,
                        'Mean K (1/cycles)': K_mean,
                        '95% CI Upper K (1/cycles)': K_upper,
                        '95% CI Lower Performance (kg CO2/kg sorbent)': Performance_lower,
                        'Mean Performance (kg CO2/kg sorbent)': Performance_mean,
                        '95% CI Upper Performance (kg CO2/kg sorbent)': Performance_upper,
                    })
    
    # Create DataFrame with all statistics
    MC_stats_df = _filter_excluded_mcs_sorbents_df(pd.DataFrame(summary_list))
    
    # Save to Excel file (create new or append to existing)
    try:
        # Try to append to existing file
        with pd.ExcelWriter('Adsorption Performance.xlsx', engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            sheet_name = f"Adsorption_Performance"
            MC_stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Saved {len(MC_stats_df)} rows to {sheet_name} sheet in Adsorption Performance.xlsx")
    except FileNotFoundError:
        # Create new file if it doesn't exist
        with pd.ExcelWriter('Adsorption Performance.xlsx', engine="openpyxl") as writer:
            sheet_name = f"Adsorption_Performance"
            MC_stats_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✅ Created new file and saved {len(MC_stats_df)} rows to {sheet_name} sheet in Adsorption Performance.xlsx")
    
    return MC_stats_df
